"""
import_toothy.py -- Mouse info and bad channels out of the Toothy spreadsheet.

The PTEN Toothy workbook is where the lab has been keeping what it knows about
each animal and each recording. Two parts of it belong in BARRY:

    mouse info      group and subgroup (PTEN / CTL, IED+ / IED-), which are
                    facts about the animal and so go on the mouse record
    bad channels    per recording, which is where BARRY already keeps them

Deliberately NOT imported: the curation columns (DS#, Garbage#, Flag#, Deep
Rev., Curation, Cur_Initials). Curation state has a home of its own with a
vocabulary and per-event provenance, and copying a tally into it would create
a set that claims to be curated without anything behind it.

Also not imported here, but available: the "Channel Data" sheet carries a
layer per channel (CA1, CA1 SP, HIL, ...) for every recording. That is a
StrataScope sheet, and --layers will bring it in.

    python tools/import_toothy.py                    # dry run
    python tools/import_toothy.py --write
    python tools/import_toothy.py --write --layers   # channel layers too
"""
from __future__ import annotations

import argparse
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
APP = os.path.dirname(HERE)
sys.path.insert(0, APP)

from backend import ids, layers as layermod, mice, sessreg, store  # noqa: E402

DEFAULT_BOOK = os.path.join(APP, "PTEN Toothy Data .xlsx")

# What each sheet is called, so a renamed tab fails loudly rather than
# silently importing nothing.
SHEET_SESSIONS = "Recording Sessions"
SHEET_TOOTHY = "Toothy Input"
SHEET_CHANNELS = "Channel Data"

# The layer names the spreadsheet uses, mapped onto StrataScope's ids. Only
# an exact match is taken: a name nobody recognises is reported rather than
# guessed at, because a wrong layer is worse than a missing one.
LAYER_MAP = {
    "CA1": "ca1_sr", "CA1 SO": "ca1_so", "CA1 SP": "ca1_sp",
    "CA1 SLM": "ca1_slm", "SLM": "ca1_slm",
    "DG OML1": "dg_oml1", "DG MML1": "dg_mml1", "DG GCL1": "dg_gcl1",
    "HIL": "hil", "HILUS": "hil",
    "DG GCL2": "dg_gcl2", "DG MML2": "dg_mml2", "DG OML2": "dg_oml2",
    "DG": "dg", "OUT": "out",
}


def cell(v):
    if v is None:
        return ""
    s = str(v).strip()
    # openpyxl hands back floats for integer-looking cells: "59.0" is 59.
    if re.fullmatch(r"-?\d+\.0", s):
        s = s[:-2]
    return s


def as_int(v):
    s = cell(v)
    try:
        return int(float(s))
    except (TypeError, ValueError):
        return None


def parse_channels(v):
    """"59", "41,59" and "[8, 41, 59]" all mean the same thing."""
    s = cell(v)
    if not s:
        return []
    out = []
    for tok in re.findall(r"\d+", s):
        n = int(tok)
        if n not in out:
            out.append(n)
    return sorted(out)


def parse_session_id(v):
    m = re.fullmatch(r"\s*m(\d+)s(\d+)\s*", cell(v), re.I)
    return (int(m.group(1)), int(m.group(2))) if m else (None, None)


def main():
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--book", default=DEFAULT_BOOK)
    ap.add_argument("--logs", default=os.path.join(APP, "GUI_logs"))
    ap.add_argument("--write", action="store_true")
    ap.add_argument("--layers", action="store_true",
                    help="also import the per-channel layers")
    args = ap.parse_args()

    import openpyxl

    if not os.path.isfile(args.book):
        print("No such workbook: " + args.book)
        return 2

    wb = openpyxl.load_workbook(args.book, data_only=True, read_only=True)
    for want in (SHEET_SESSIONS, SHEET_TOOTHY):
        if want not in wb.sheetnames:
            print("The workbook has no %r sheet. Found: %s"
                  % (want, ", ".join(wb.sheetnames)))
            return 2

    st = store.Store(args.logs, auto_stage=False)
    reg = sessreg.Registry(st)
    book = mice.MouseBook(args.logs, st)
    known = st.all_sessions()

    print("Reading  %s" % args.book)
    print("Mode     %s" % ("WRITE" if args.write else "dry run"))
    print()

    # ---- mouse info -------------------------------------------------------
    def rows_of(sheet):
        it = wb[sheet].iter_rows(values_only=True)
        head = [cell(h).strip().lower() for h in next(it)]
        for raw in it:
            if not any(x is not None and str(x).strip() for x in raw):
                continue
            yield dict(zip(head, raw))

    per_mouse = {}
    bad_by_session = {}
    conditions = {}

    for r in rows_of(SHEET_SESSIONS):
        mouse = as_int(r.get("mouse_id") if r.get("mouse_id") is not None
                       else re.sub(r"\D", "", cell(r.get("mouse_id"))))
        if mouse is None:
            m = re.fullmatch(r"m(\d+)", cell(r.get("mouse_id")), re.I)
            mouse = int(m.group(1)) if m else None
        if mouse is None:
            continue
        grp, sub = cell(r.get("group")), cell(r.get("subgroup"))
        slot = per_mouse.setdefault(mouse, {})
        if grp:
            slot["group"] = grp
        if sub:
            slot["subgroup"] = sub

        mo, se = parse_session_id(r.get("session id"))
        if mo is not None:
            chans = parse_channels(r.get("bad channel"))
            if chans:
                bad_by_session[(mo, se)] = chans
            cond = cell(r.get("condition"))
            if cond:
                conditions[(mo, se)] = cond

    # Toothy Input carries the same bad channels, sometimes more complete.
    for r in rows_of(SHEET_TOOTHY):
        mo, se = parse_session_id(r.get("session id"))
        if mo is None:
            continue
        chans = parse_channels(r.get("bad channels"))
        if chans:
            have = set(bad_by_session.get((mo, se), []))
            bad_by_session[(mo, se)] = sorted(have | set(chans))

    print("%d mice with info, %d sessions with bad channels"
          % (len(per_mouse), len(bad_by_session)))

    # ---- write the mice ---------------------------------------------------
    # Which project each mouse is in, taken from the registry rather than
    # guessed: the spreadsheet's "group" is the experimental group (PTEN vs
    # CTL), which is not the same thing as the project.
    proj_of = {}
    for rec in known:
        m = rec.get("mouse")
        if m is not None:
            proj_of.setdefault(m, rec.get("project") or "Unfiled")

    wrote_mice = 0
    unknown_mice = []
    for mouse, attrs in sorted(per_mouse.items()):
        project = proj_of.get(mouse)
        if not project:
            unknown_mice.append(mouse)
            continue
        if args.write:
            book.set(project, mouse, attrs)
        wrote_mice += 1

    # ---- write the bad channels ------------------------------------------
    wrote_bad = 0
    unmatched = []
    for (mouse, session), chans in sorted(bad_by_session.items()):
        ident = {
            "mouse": mouse, "session": session,
            "loose_key": ids.make_loose_key(mouse, session),
            "key": None, "start": None,
        }
        rec, how = ids.match(ident, known)
        if not rec:
            unmatched.append((mouse, session, chans))
            continue
        if args.write:
            have = set(rec.get("bad_channels") or [])
            st.set_bad_channels(
                {k: rec.get(k) for k in
                 ("key", "loose_key", "mouse", "session", "start", "label")},
                sorted(have | set(chans)),
                note="Bad channels from the Toothy workbook.")
            cond = conditions.get((mouse, session))
            if cond:
                st.upsert_session(
                    {k: rec.get(k) for k in
                     ("key", "loose_key", "mouse", "session", "start", "label")},
                    {"condition": cond})
        wrote_bad += 1

    # ---- optional: the channel layers ------------------------------------
    wrote_layers = 0
    bad_layer_names = {}
    if args.layers and SHEET_CHANNELS in wb.sheetnames:
        lay = layermod.Layers(args.logs, st)
        by_sess = {}
        for r in rows_of(SHEET_CHANNELS):
            mo, se = parse_session_id(r.get("sess"))
            if mo is None:
                continue
            num = as_int(r.get("eegnum"))
            name = cell(r.get("location")).upper()
            if num is None or not name:
                continue
            lid = LAYER_MAP.get(name)
            if not lid:
                bad_layer_names[name] = bad_layer_names.get(name, 0) + 1
                continue
            by_sess.setdefault((mo, se), {})[num] = lid

        for (mouse, session), mapping in sorted(by_sess.items()):
            ident = {"mouse": mouse, "session": session,
                     "loose_key": ids.make_loose_key(mouse, session),
                     "key": None, "start": None}
            rec, _ = ids.match(ident, known)
            if not rec or not rec.get("gid"):
                continue
            if args.write:
                lay.ensure(rec["gid"], session_label=rec.get("label"),
                           channels=sorted(mapping))
                lay.set_many(rec["gid"], mapping)
            wrote_layers += 1

    # ---- report -----------------------------------------------------------
    print()
    print("mouse records   %d" % wrote_mice)
    print("bad channels    %d session(s)" % wrote_bad)
    if args.layers:
        print("layer sheets    %d session(s)" % wrote_layers)
        if bad_layer_names:
            print("  layer names nobody recognised (left alone):")
            for k, v in sorted(bad_layer_names.items(), key=lambda kv: -kv[1]):
                print("     %-12s %d channel(s)" % (k, v))
    if unknown_mice:
        print()
        print("Mice with no registered recording, so no project to file them "
              "under: " + ", ".join("m%d" % m for m in unknown_mice))
    if unmatched:
        print()
        print("Bad channels for recordings BARRY has not met:")
        for mouse, session, chans in unmatched[:20]:
            print("   m%-4d s%-4d %s" % (mouse, session, chans))
        if len(unmatched) > 20:
            print("   ... and %d more" % (len(unmatched) - 20))
    if not args.write:
        print()
        print("Nothing written. Re-run with --write.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
